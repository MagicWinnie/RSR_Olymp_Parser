# -*- coding: utf-8 -*-
import argparse
import requests
import pandas as pd
from bs4 import BeautifulSoup


parser = argparse.ArgumentParser()
parser.add_argument("-o", "--output", type=str, required=True, help="Название выходного excel файла")
parser.add_argument("-l", "--level", type=int, help="Фильтрация по уровню РСОШ (указанный и выше)")
parser.add_argument("-w", "--whitelist", type=str, help="Путь к файлу с названиями профилей или предметов, которые нужно оставить")
parser.add_argument("-b", "--blacklist", type=str, help="Путь к файлу с названиями профилей или предметов, которые нужно убрать")
parser.add_argument("-c", "--concat", type=bool, default=True, help="Объединять ли последовательно идущие ячейки с одинаковыми значениями")
parser.add_argument("-s", "--style", type=bool, default=True, help="Форматировать ли таблицу")
args = parser.parse_args()


req_headers = {
    'Host': 'rsr-olymp.ru',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
    'Accept-Encoding': 'gzip, deflate',
    'Accept-Language': 'ru,en-US;q=0.9,en;q=0.8'
}

DATAURL = 'http://rsr-olymp.ru/'

### GETTING THE TABLE
session = requests.session()

html_rsr = session.get(DATAURL, headers=req_headers, allow_redirects=True).text
soup = BeautifulSoup(html_rsr, features="lxml")
html_table = soup.find_all('table')[-1]
###

### GETTING URLS
links = html_table.find_all('a', href = True)
links = [x['href'] for x in links]
for i in range(len(links)):
    if 'http://' in links[i]:
        links[i] = links[i].replace('http://', 'https://')
    if 'http' not in links[i]:
        links[i] = 'https://' + links[i]
###

### PARSING THE TABLE
df = pd.read_html(str(html_table), index_col=False)[0]
###

### ADDING URLS TO THE TABLE
links_out = []
frequency = df['Название'].value_counts()
visited = set()
j = 0
for i in range(len(df['Название'])):
    if df['Название'][i] not in visited:
        links_out += [links[j]] * frequency[df['Название'][i]]
        visited.add(df['Название'][i])
        j += 1

df['Ссылка'] = links_out
###

### FILTERING
if args.level is not None:
    assert (args.level >= df['Уровень'].min() and args.level <= df['Уровень'].max()), "Введен некорректный уровень"
    df = df.loc[df['Уровень'] <= args.level]
df = df.reset_index(drop=True)
if args.whitelist is not None:
    with open(args.whitelist, 'r', encoding='utf-8') as f:
        whitelist = f.readlines()
    whitelist = list(map(lambda x: x.lower().replace('\n', ''), whitelist))
    
    df_temp = pd.DataFrame(columns=list(df))
    for i in range(len(df['Предметы'])):
        temp_sub = list(map(str.lower, df['Предметы'][i].replace(' и ', ', ').split(', ')))
        temp_pro = list(map(str.lower, df['Профиль'][i].replace(' и ', ', ').split(', ')))
        if (len(set(temp_sub).intersection(whitelist)) > 0) or (len(set(temp_pro).intersection(whitelist)) > 0):
            df_temp.loc[len(df_temp)] = df.iloc[i]
    df = df_temp
df = df.reset_index(drop=True)
if args.blacklist is not None:
    with open(args.blacklist, 'r') as f:
        blacklist = f.readlines()
    blacklist = list(map(str.lower, blacklist))
    
    df_temp = pd.DataFrame(columns=list(df))
    for i in range(len(df['Предметы'])):
        temp_sub = list(map(str.lower, df['Предметы'][i].replace(' и ', ', ').split(', ')))
        temp_pro = list(map(str.lower, df['Профиль'][i].replace(' и ', ', ').split(', ')))
        if (len(set(temp_sub).intersection(blacklist)) == 0) or (len(set(temp_pro).intersection(blacklist)) == 0):
            df_temp.loc[len(df_temp)] = df.iloc[i]
    df = df_temp
df = df.reset_index(drop=True)
###

### OUTPUT
df.to_excel(args.output, index=False, sheet_name='Sheet1')
###

### MERGING CELLS WITH THE SAME VALUE
if args.concat:
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
    except ModuleNotFoundError:
        print("openpyxl не найден!\nВыполните 'pip install openpyxl' чтобы установить.\nНе объединяю ячейки.")
        exit(-1)

    ranges = []
    visited = set()
    for i in range(len(df["Название"])):
        if df['Название'][i] not in visited:
            ranges.append(i)
            visited.add(df['Название'][i])
    ranges.append(len(df['Название']))
    ranges = [x + 2 for x in ranges]

    wb = load_workbook(filename = args.output)
    ws = wb['Sheet1']

    for i in range(len(ranges) - 1):
        ws.merge_cells(start_row=ranges[i], start_column=1, end_row=ranges[i + 1] - 1, end_column=1)
        ws.merge_cells(start_row=ranges[i], start_column=2, end_row=ranges[i + 1] - 1, end_column=2)
        ws.merge_cells(start_row=ranges[i], start_column=6, end_row=ranges[i + 1] - 1, end_column=6)
    
    wb.save(args.output)
###
### STYLE FORMATTING
if args.style:
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, NamedStyle
    except ModuleNotFoundError:
        print("openpyxl не найден!\nВыполните 'pip install openpyxl' чтобы установить.\nНе объединяю ячейки.")
        exit(-1)

    ### FINDING COORDINATES OF CELLS TO MERGE
    ranges = []
    visited = set()
    for i in range(len(df["Название"])):
        if df['Название'][i] not in visited:
            ranges.append(i)
            visited.add(df['Название'][i])
    ranges.append(len(df['Название']))
    ranges = [x + 2 for x in ranges]
    ###

    wb = load_workbook(filename = args.output)
    ws = wb['Sheet1']

    ### STYLE FOR ALL CELLS
    all_cells = NamedStyle(name="all_cells")
    bd = Side(style='thin', color="000000")
    all_cells.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    all_cells.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wb.add_named_style(all_cells)
    ###

    for l in ['A', 'B', 'C', 'D', 'E', 'F']:
        for cell in ws[l]:
            cell.style = 'all_cells'

    for l in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws[l + '1'].font = Font(bold=True)

    ### APPLYING CUSTOM COLUMN WIDTH
    dims = {
        'A': 40,
        'B': 450,
        'C': 250,
        'D': 400,
        'E': 82,
        'F': 300
    }
    for col, value in dims.items():
        ws.column_dimensions[col].width = value/10
    ###

    wb.save(args.output)
###
